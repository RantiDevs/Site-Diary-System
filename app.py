from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import json
import os
import io
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or os.urandom(24).hex()

database_url = os.environ.get('NEON_DATABASE_URL') or os.environ.get('DATABASE_URL', 'sqlite:///site_diary.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

db = SQLAlchemy(app)


class DiaryEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contract_name = db.Column(db.String(200), default='LADOL POWER STATION')
    contract_no = db.Column(db.String(50), default='LPS-003')
    site = db.Column(db.String(200), default='LADOL Power')
    entry_date = db.Column(db.Date, nullable=False)
    day_of_week = db.Column(db.String(20))
    shift = db.Column(db.String(20), default='Day Shift')
    pm = db.Column(db.String(100))
    pe = db.Column(db.String(200))
    project_engineer = db.Column(db.String(200))
    engineers = db.Column(db.String(300))
    foreman = db.Column(db.String(100))
    hour_from = db.Column(db.String(20), default='8:00AM')
    hour_to = db.Column(db.String(20), default='5:00PM')
    weather = db.Column(db.String(100))
    activities = db.Column(db.Text)
    others = db.Column(db.Text)
    note = db.Column(db.Text)
    deliveries = db.Column(db.Text)
    collections = db.Column(db.Text)
    labour = db.Column(db.Text)
    plant = db.Column(db.Text)
    prepared_by = db.Column(db.String(100))
    prepared_date = db.Column(db.Date)
    accepted_by = db.Column(db.String(100))
    accepted_date = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def get_activities(self):
        return json.loads(self.activities) if self.activities else []

    def get_deliveries(self):
        return json.loads(self.deliveries) if self.deliveries else []

    def get_collections(self):
        return json.loads(self.collections) if self.collections else []

    def get_labour(self):
        return json.loads(self.labour) if self.labour else {}

    def get_plant(self):
        return json.loads(self.plant) if self.plant else {}


LABOUR_CATEGORIES = [
    'Engineers', 'QS', 'HSE', 'NYSC', 'I.T',
    'Welders and Steel Fabricators', 'Welder Helper',
    'Store Keeper', 'Skilled Labour', 'Operators',
    'Carpenter', 'Iron Benders', 'Scaffolders',
    'Painters', 'Weld Inspectors',
    'Equipment Technician/Mechanic', 'Mason',
    'Machine Operators', 'Rigger', 'Site Helpers'
]

PLANT_CATEGORIES = [
    '114Kva Generator', 'Mobile Crane', 'Fork Lift',
    'Scissors Lift', 'Concrete Mixer',
    'Hyundai HL660L Wheel Loader (17.3 ton)',
    'Bobcat S130 Skid-Steer Loader (2.37 ton)',
    'Staunch RL6000 Drum Roller (600kg)',
    'Telescopic Crane', 'AGR Payloader'
]

DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
SHIFTS = ['Day Shift', 'Night Shift']


@app.route('/')
def index():
    entries = DiaryEntry.query.order_by(DiaryEntry.entry_date.desc()).all()
    return render_template('index.html', entries=entries)


@app.route('/new', methods=['GET', 'POST'])
def new_entry():
    if request.method == 'POST':
        try:
            entry_date = datetime.strptime(request.form['entry_date'], '%Y-%m-%d').date()
            day_of_week = entry_date.strftime('%A')

            activities = []
            act_texts = request.form.getlist('activity_text[]')
            act_delays = request.form.getlist('activity_delay[]')
            for t, d in zip(act_texts, act_delays):
                if t.strip():
                    activities.append({'text': t.strip(), 'delay': d.strip()})

            deliveries = [d.strip() for d in request.form.getlist('delivery[]') if d.strip()]
            collections = [c.strip() for c in request.form.getlist('collection[]') if c.strip()]

            labour = {}
            for cat in LABOUR_CATEGORIES:
                val = request.form.get(f'labour_{cat}', '0')
                if val and val.strip():
                    labour[cat] = int(val) if val.isdigit() else 0

            plant = {}
            for p in PLANT_CATEGORIES:
                val = request.form.get(f'plant_{p}', '0')
                if val and val.strip():
                    plant[p] = int(val) if val.isdigit() else 0

            custom_labour_names = request.form.getlist('custom_labour_name[]')
            custom_labour_counts = request.form.getlist('custom_labour_count[]')
            for name, count in zip(custom_labour_names, custom_labour_counts):
                if name.strip():
                    labour[name.strip()] = int(count) if count.isdigit() else 0

            custom_plant_names = request.form.getlist('custom_plant_name[]')
            custom_plant_counts = request.form.getlist('custom_plant_count[]')
            for name, count in zip(custom_plant_names, custom_plant_counts):
                if name.strip():
                    plant[name.strip()] = int(count) if count.isdigit() else 0

            prepared_date_str = request.form.get('prepared_date')
            accepted_date_str = request.form.get('accepted_date')
            prepared_date = datetime.strptime(prepared_date_str, '%Y-%m-%d').date() if prepared_date_str else None
            accepted_date = datetime.strptime(accepted_date_str, '%Y-%m-%d').date() if accepted_date_str else None

            from_h = request.form.get('hour_from_h', '8')
            from_m = request.form.get('hour_from_m', '00')
            from_p = request.form.get('hour_from_p', 'AM')
            to_h = request.form.get('hour_to_h', '5')
            to_m = request.form.get('hour_to_m', '00')
            to_p = request.form.get('hour_to_p', 'PM')
            hour_from = f"{from_h}:{from_m}{from_p}"
            hour_to = f"{to_h}:{to_m}{to_p}"

            entry = DiaryEntry(
                contract_name=request.form.get('contract_name', 'LADOL POWER STATION'),
                contract_no=request.form.get('contract_no', 'LPS-003'),
                site=request.form.get('site', 'LADOL Power'),
                entry_date=entry_date,
                day_of_week=day_of_week,
                shift=request.form.get('shift', 'Day Shift'),
                pm=request.form.get('pm', ''),
                pe=request.form.get('pe', ''),
                project_engineer=request.form.get('project_engineer', ''),
                engineers=request.form.get('engineers', ''),
                foreman=request.form.get('foreman', ''),
                hour_from=hour_from,
                hour_to=hour_to,
                weather=request.form.get('weather', ''),
                activities=json.dumps(activities),
                others=request.form.get('others', ''),
                note=request.form.get('note', ''),
                deliveries=json.dumps(deliveries),
                collections=json.dumps(collections),
                labour=json.dumps(labour),
                plant=json.dumps(plant),
                prepared_by=request.form.get('prepared_by', ''),
                prepared_date=prepared_date,
                accepted_by=request.form.get('accepted_by', ''),
                accepted_date=accepted_date
            )
            db.session.add(entry)
            db.session.commit()
            flash('Diary entry saved successfully!', 'success')
            return redirect(url_for('view_entry', entry_id=entry.id))
        except Exception as e:
            flash(f'Error saving entry: {str(e)}', 'error')

    today = date.today()
    return render_template('form.html',
                           entry=None,
                           labour_categories=LABOUR_CATEGORIES,
                           plant_categories=PLANT_CATEGORIES,
                           days=DAYS_OF_WEEK,
                           shifts=SHIFTS,
                           today=today.strftime('%Y-%m-%d'))


@app.route('/view/<int:entry_id>')
def view_entry(entry_id):
    entry = db.get_or_404(DiaryEntry, entry_id)
    return render_template('view.html', entry=entry,
                           labour_categories=LABOUR_CATEGORIES,
                           plant_categories=PLANT_CATEGORIES)


@app.route('/edit/<int:entry_id>', methods=['GET', 'POST'])
def edit_entry(entry_id):
    entry = db.get_or_404(DiaryEntry, entry_id)
    if request.method == 'POST':
        try:
            entry_date = datetime.strptime(request.form['entry_date'], '%Y-%m-%d').date()
            day_of_week = entry_date.strftime('%A')

            activities = []
            act_texts = request.form.getlist('activity_text[]')
            act_delays = request.form.getlist('activity_delay[]')
            for t, d in zip(act_texts, act_delays):
                if t.strip():
                    activities.append({'text': t.strip(), 'delay': d.strip()})

            deliveries = [d.strip() for d in request.form.getlist('delivery[]') if d.strip()]
            collections = [c.strip() for c in request.form.getlist('collection[]') if c.strip()]

            labour = {}
            for cat in LABOUR_CATEGORIES:
                val = request.form.get(f'labour_{cat}', '0')
                if val and val.strip():
                    labour[cat] = int(val) if val.isdigit() else 0

            plant = {}
            for p in PLANT_CATEGORIES:
                val = request.form.get(f'plant_{p}', '0')
                if val and val.strip():
                    plant[p] = int(val) if val.isdigit() else 0

            custom_labour_names = request.form.getlist('custom_labour_name[]')
            custom_labour_counts = request.form.getlist('custom_labour_count[]')
            for name, count in zip(custom_labour_names, custom_labour_counts):
                if name.strip():
                    labour[name.strip()] = int(count) if count.isdigit() else 0

            custom_plant_names = request.form.getlist('custom_plant_name[]')
            custom_plant_counts = request.form.getlist('custom_plant_count[]')
            for name, count in zip(custom_plant_names, custom_plant_counts):
                if name.strip():
                    plant[name.strip()] = int(count) if count.isdigit() else 0

            prepared_date_str = request.form.get('prepared_date')
            accepted_date_str = request.form.get('accepted_date')
            prepared_date = datetime.strptime(prepared_date_str, '%Y-%m-%d').date() if prepared_date_str else None
            accepted_date = datetime.strptime(accepted_date_str, '%Y-%m-%d').date() if accepted_date_str else None

            entry.contract_name = request.form.get('contract_name', entry.contract_name)
            entry.contract_no = request.form.get('contract_no', entry.contract_no)
            entry.site = request.form.get('site', entry.site)
            entry.entry_date = entry_date
            entry.day_of_week = day_of_week
            entry.shift = request.form.get('shift', entry.shift)
            entry.pm = request.form.get('pm', '')
            entry.pe = request.form.get('pe', '')
            entry.project_engineer = request.form.get('project_engineer', '')
            entry.engineers = request.form.get('engineers', '')
            entry.foreman = request.form.get('foreman', '')
            from_h = request.form.get('hour_from_h', '8')
            from_m = request.form.get('hour_from_m', '00')
            from_p = request.form.get('hour_from_p', 'AM')
            to_h = request.form.get('hour_to_h', '5')
            to_m = request.form.get('hour_to_m', '00')
            to_p = request.form.get('hour_to_p', 'PM')
            entry.hour_from = f"{from_h}:{from_m}{from_p}"
            entry.hour_to = f"{to_h}:{to_m}{to_p}"
            entry.weather = request.form.get('weather', '')
            entry.activities = json.dumps(activities)
            entry.others = request.form.get('others', '')
            entry.note = request.form.get('note', '')
            entry.deliveries = json.dumps(deliveries)
            entry.collections = json.dumps(collections)
            entry.labour = json.dumps(labour)
            entry.plant = json.dumps(plant)
            entry.prepared_by = request.form.get('prepared_by', '')
            entry.prepared_date = prepared_date
            entry.accepted_by = request.form.get('accepted_by', '')
            entry.accepted_date = accepted_date

            db.session.commit()
            flash('Entry updated successfully!', 'success')
            return redirect(url_for('view_entry', entry_id=entry.id))
        except Exception as e:
            flash(f'Error updating entry: {str(e)}', 'error')

    return render_template('form.html',
                           entry=entry,
                           labour_categories=LABOUR_CATEGORIES,
                           plant_categories=PLANT_CATEGORIES,
                           days=DAYS_OF_WEEK,
                           shifts=SHIFTS,
                           today=entry.entry_date.strftime('%Y-%m-%d') if entry.entry_date else '')


@app.route('/delete/<int:entry_id>', methods=['POST'])
def delete_entry(entry_id):
    entry = db.get_or_404(DiaryEntry, entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash('Entry deleted.', 'info')
    return redirect(url_for('index'))


@app.route('/print/<int:entry_id>')
def print_entry(entry_id):
    entry = db.get_or_404(DiaryEntry, entry_id)
    return render_template('print.html', entry=entry,
                           labour_categories=LABOUR_CATEGORIES,
                           plant_categories=PLANT_CATEGORIES)


@app.route('/export/<int:entry_id>')
def export_excel(entry_id):
    entry = db.get_or_404(DiaryEntry, entry_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entry.entry_date.strftime('%A %d %b %Y')

    def border(style='thin'):
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)

    def cell(row, col, value='', bold=False, size=10, bg=None, fg='000000',
             align='left', valign='center', wrap=False, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=fg, italic=italic, name='Calibri')
        c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        c.border = border()
        return c

    NAVY   = '1A3C5E'
    ACCENT = 'E8941A'
    GREEN  = '93C47D'
    PINK   = 'EA9999'
    LBLUE  = 'CFE2F3'
    YELLOW = 'FFE599'
    LGREY  = 'F3F3F3'
    RED    = 'CC0000'

    r = 1

    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 16

    ws.merge_cells('D1:M1')
    c = ws.cell(row=1, column=4, value='AGR Construction')
    c.font = Font(bold=True, size=16, color=NAVY, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='bottom')

    ws.merge_cells('D2:M2')
    c = ws.cell(row=2, column=4,
                value='Design | Construction | Maintenance  |  RC 718729')
    c.font = Font(italic=True, size=9, color='444444', name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='top')

    header_border = Border(bottom=Side(style='medium', color=NAVY))
    for col in range(1, 14):
        ws.cell(row=2, column=col).border = header_border

    logo_path = os.path.join(app.root_path, 'static', 'images', 'agr_logo.jpg')
    if os.path.exists(logo_path):
        try:
            logo_img = XLImage(logo_path)
            logo_img.height = 52
            logo_img.width = 68
            ws.add_image(logo_img, 'A1')
        except Exception:
            pass

    r = 3

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'Contract name:', bold=True, bg=LGREY)
    ws.merge_cells(f'B{r}:H{r}')
    cell(r, 2, entry.contract_name, bold=True)
    cell(r, 9, 'Contract no:', bold=True, bg=LGREY, align='right')
    ws.merge_cells(f'J{r}:M{r}')
    cell(r, 10, entry.contract_no, bold=True)
    r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'Site:', bold=True, bg=LGREY)
    ws.merge_cells(f'B{r}:D{r}')
    cell(r, 2, entry.site)
    cell(r, 5, entry.day_of_week or '', bold=True)
    ws.merge_cells(f'F{r}:H{r}')
    cell(r, 6, entry.entry_date.strftime('%d/%m/%Y'), bold=True, fg='0000CC')
    ws.merge_cells(f'J{r}:M{r}')
    cell(r, 10, entry.shift, bold=True, bg=YELLOW, align='center')
    r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'PD:', bold=True, bg=LGREY)
    ws.merge_cells(f'B{r}:D{r}')
    cell(r, 2, entry.pe or '')
    cell(r, 5, 'PM:', bold=True, bg=LGREY)
    ws.merge_cells(f'F{r}:H{r}')
    cell(r, 6, entry.pm or '')
    cell(r, 9, 'Foreman:', bold=True, bg=LGREY)
    ws.merge_cells(f'J{r}:M{r}')
    cell(r, 10, entry.foreman or '')
    r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'PE:', bold=True, bg=LGREY)
    ws.merge_cells(f'B{r}:D{r}')
    cell(r, 2, entry.project_engineer or '')
    cell(r, 5, 'Engineers:', bold=True, bg=LGREY)
    ws.merge_cells(f'F{r}:H{r}')
    cell(r, 6, entry.engineers or '')
    r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'Working hours:', bold=True, bg=LGREY)
    cell(r, 2, 'from', align='center')
    cell(r, 3, fmt_time(entry.hour_from), bold=True, align='center')
    cell(r, 4, 'to', align='center')
    cell(r, 5, fmt_time(entry.hour_to), bold=True, align='center')
    cell(r, 9, 'Temperature:', bold=True, bg=LGREY, align='right')
    ws.merge_cells(f'J{r}:M{r}')
    cell(r, 10, fmt_weather(entry.weather), bold=True, align='center')
    r += 1

    ws.row_dimensions[r].height = 18
    ws.merge_cells(f'A{r}:M{r}')
    c = ws.cell(row=r, column=1, value='Activities')
    c.font = Font(bold=True, size=11, color='FFFFFF', name='Calibri')
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    ws.row_dimensions[r].height = 16
    cell(r, 1, '#', bold=True, bg=LGREY, align='center')
    ws.merge_cells(f'B{r}:H{r}')
    cell(r, 2, 'Actual Work Done', bold=True, bg=LGREY)
    ws.merge_cells(f'I{r}:M{r}')
    cell(r, 9, 'Delays:', bold=True, bg=LGREY)
    r += 1

    acts = entry.get_activities()
    for i, act in enumerate(acts, 1):
        ws.row_dimensions[r].height = max(18, min(len(act['text']) // 6, 60))
        cell(r, 1, i, align='center')
        ws.merge_cells(f'B{r}:H{r}')
        cell(r, 2, act['text'], wrap=True)
        ws.merge_cells(f'I{r}:M{r}')
        delay = act.get('delay', 'Nil') or 'Nil'
        delay_bg = YELLOW if delay and delay != 'Nil' else None
        cell(r, 9, delay, align='center', bg=delay_bg)
        r += 1

    if not acts:
        ws.merge_cells(f'B{r}:H{r}')
        cell(r, 2, '')
        ws.merge_cells(f'I{r}:M{r}')
        cell(r, 9, '')
        r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'Others', bold=True, bg=GREEN)
    ws.merge_cells(f'B{r}:M{r}')
    cell(r, 2, entry.others or '', bg='EAF5E0', wrap=True)
    r += 1

    ws.row_dimensions[r].height = 20
    cell(r, 1, 'Note', bold=True, bg=PINK)
    ws.merge_cells(f'B{r}:M{r}')
    cell(r, 2, entry.note or '', bg='FDF0F0', wrap=True)
    r += 1

    ws.row_dimensions[r].height = 16
    cell(r, 1, '#', bold=True, bg=LBLUE, align='center')
    ws.merge_cells(f'B{r}:G{r}')
    cell(r, 2, 'Deliveries:', bold=True, bg=LBLUE)
    cell(r, 8, '#', bold=True, bg=LBLUE, align='center')
    ws.merge_cells(f'I{r}:M{r}')
    cell(r, 9, 'Collections:', bold=True, bg=LBLUE)
    r += 1

    deliveries = entry.get_deliveries()
    collections = entry.get_collections()
    max_dc = max(len(deliveries), len(collections), 1)
    for i in range(max_dc):
        ws.row_dimensions[r].height = 16
        cell(r, 1, i + 1, align='center')
        ws.merge_cells(f'B{r}:G{r}')
        cell(r, 2, deliveries[i] if i < len(deliveries) else '', wrap=True)
        cell(r, 8, i + 1, align='center')
        ws.merge_cells(f'I{r}:M{r}')
        cell(r, 9, collections[i] if i < len(collections) else '')
        r += 1

    ws.row_dimensions[r].height = 18
    lab = entry.get_labour()
    total_workers = sum(lab.values())
    cell(r, 1, 'Labour resource:', bold=True, bg=GREEN)
    ws.merge_cells(f'B{r}:D{r}')
    cell(r, 2, 'Total:', bold=True, bg=GREEN, align='right')
    cell(r, 5, total_workers, bold=True, fg=RED, bg=GREEN, align='center', size=12)
    ws.merge_cells(f'F{r}:G{r}')
    cell(r, 6, '', bg=GREEN)
    cell(r, 8, 'Plant:', bold=True, bg=PINK)
    ws.merge_cells(f'I{r}:M{r}')
    cell(r, 9, '', bg=PINK)
    r += 1

    plt = entry.get_plant()
    lab_items = list(lab.items())
    plt_items = [(k, v) for k, v in plt.items() if v > 0]
    max_lp = max(len(lab_items), len(plt_items), 1)

    for i in range(max_lp):
        ws.row_dimensions[r].height = 16
        if i < len(lab_items):
            lname, lcount = lab_items[i]
            ws.merge_cells(f'A{r}:D{r}')
            cell(r, 1, lname)
            cell(r, 5, lcount if lcount else 0, align='center')
            ws.merge_cells(f'F{r}:G{r}')
            cell(r, 6, '')
        else:
            ws.merge_cells(f'A{r}:G{r}')
            cell(r, 1, '')

        if i < len(plt_items):
            pname, pcount = plt_items[i]
            ws.merge_cells(f'H{r}:L{r}')
            cell(r, 8, pname)
            cell(r, 13, pcount if pcount else 0, align='center')
        else:
            ws.merge_cells(f'H{r}:M{r}')
            cell(r, 8, '')
        r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'Prepared by:', bold=True, bg=LGREY)
    ws.merge_cells(f'B{r}:D{r}')
    cell(r, 2, entry.prepared_by or '')
    cell(r, 5, 'Signed:', bold=True, bg=LGREY)
    ws.merge_cells(f'F{r}:H{r}')
    cell(r, 6, '')
    cell(r, 9, 'Date:', bold=True, bg=LGREY, align='right')
    ws.merge_cells(f'J{r}:M{r}')
    cell(r, 10, entry.prepared_date.strftime('%d/%m/%Y') if entry.prepared_date else '')
    r += 1

    ws.row_dimensions[r].height = 18
    cell(r, 1, 'Accepted by:', bold=True, bg=LGREY)
    ws.merge_cells(f'B{r}:D{r}')
    cell(r, 2, entry.accepted_by or '')
    cell(r, 5, 'Signed:', bold=True, bg=LGREY)
    ws.merge_cells(f'F{r}:H{r}')
    cell(r, 6, '')
    cell(r, 9, 'Date:', bold=True, bg=LGREY, align='right')
    ws.merge_cells(f'J{r}:M{r}')
    cell(r, 10, entry.accepted_date.strftime('%d/%m/%Y') if entry.accepted_date else '')
    r += 1

    col_widths = {
        1: 18, 2: 14, 3: 10, 4: 10, 5: 10,
        6: 14, 7: 8, 8: 8, 9: 22, 10: 10,
        11: 8, 12: 8, 13: 8,
    }
    for col_idx, default_w in col_widths.items():
        col_letter = get_column_letter(col_idx)
        max_len = default_w
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row_cells:
                try:
                    if c.value:
                        max_len = max(max_len, len(str(c.value)) + 2)
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len, 40)

    ws.freeze_panes = 'A3'
    ws.sheet_view.showGridLines = True
    ws.print_title_rows = '1:2'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (f"SiteDiary_{entry.entry_date.strftime('%d%m%Y')}"
                f"_{entry.shift.replace(' ', '')}.xlsx")
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def fmt_time(t):
    if not t:
        return '—'
    t = str(t).strip()
    try:
        upper = t.upper()
        if upper.endswith('AM') or upper.endswith('PM'):
            suffix = upper[-2:]
            parts = upper[:-2].split(':')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            if suffix == 'PM' and h != 12:
                h += 12
            elif suffix == 'AM' and h == 12:
                h = 0
        else:
            parts = t.split(':')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
        out_suffix = 'AM' if h < 12 else 'PM'
        h12 = h % 12 or 12
        return f'{h12}:{m:02d} {out_suffix}'
    except Exception:
        return t


def parse_time(t):
    if not t:
        return (8, '00', 'AM')
    t = str(t).strip()
    try:
        upper = t.upper()
        if upper.endswith('AM') or upper.endswith('PM'):
            period = upper[-2:]
            parts = upper[:-2].split(':')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
        else:
            parts = t.split(':')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            period = 'AM' if h < 12 else 'PM'
            h = h % 12 or 12
        return (h, f'{m:02d}', period)
    except Exception:
        return (8, '00', 'AM')


app.jinja_env.globals['fmt_time'] = fmt_time
app.jinja_env.globals['parse_time'] = parse_time


def fmt_weather(w):
    if not w:
        return '—'
    w = str(w).strip()
    if w and not w.endswith('°C') and not any(c.isalpha() for c in w):
        return f'{w}°C'
    return w

app.jinja_env.globals['fmt_weather'] = fmt_weather

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
